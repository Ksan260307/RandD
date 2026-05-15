import os
import textwrap
import torch
import warnings
import pandas as pd
import openpyxl
from transformers import AutoModelForCausalLM, AutoTokenizer, TextStreamer
from transformers import logging as hf_logging

# 警告メッセージとHugging Faceのログを完全に沈黙させる
warnings.filterwarnings("ignore")
os.environ["TRANSFORMERS_VERBOSITY"] = "error"
os.environ["TOKENIZERS_PARALLELISM"] = "false"
hf_logging.set_verbosity_error()

# =====================================================================
# ハードウェア極限最適化 (PyTorch)
# =====================================================================
torch.set_num_threads(os.cpu_count() or 4)

# 【計算力向上】非正規数(Denormal)によるCPUの深刻な計算遅延ペナルティをハードウェアレベルで回避
torch.set_flush_denormal(True)

if torch.cuda.is_available():
    torch.backends.cuda.matmul.allow_tf32 = True
    torch.backends.cudnn.allow_tf32 = True
    torch.backends.cudnn.benchmark = True

# =====================================================================
# 内部ロジック: Py_UCD-F_ABC アーキテクチャベースのコンテキスト管理エンジン
# =====================================================================

class PyUCDF_ContextEngine:
    def __init__(self, sector_capacity=5, max_sectors=2000):
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        self.sector_capacity = sector_capacity
        self.max_sectors = max_sectors
        
        self.ucdf_sectors = []
        self.probability_cloud_summary = ""
        self.conversation_turns = []
        
        self.vocab_size = 65536 
        
        self.state_buffer = torch.zeros(max_sectors, dtype=torch.int32, device=self.device)
        self.feature_matrix = torch.zeros((max_sectors, self.vocab_size), dtype=torch.int16, device=self.device)
        self.document_frequency = torch.zeros(self.vocab_size, dtype=torch.int16, device=self.device)
        
        # 検索時の `torch.sum` 計算を排除するための事前計算キャッシュ（SoA構造）
        self.sector_lengths = torch.ones(max_sectors, dtype=torch.int32, device=self.device)
        
        self.active_sector_count = 0

    def _hash_bigram(self, bigram):
        return hash(bigram) % self.vocab_size

    def pack_state(self, idx, va, vb, vc, state, ruin):
        packed = ((int(va) & 0xFF) | 
                  ((int(vb) & 0xFF) << 8) | 
                  ((int(vc) & 0xFF) << 16) | 
                  ((int(state) & 0x07) << 24) | 
                  ((int(ruin) & 0x07) << 27))
        self.state_buffer[idx] = torch.tensor(packed, dtype=torch.int32, device=self.device)

    def unpack_state(self, idx):
        val = self.state_buffer[idx].item()
        return {
            "vA": val & 0xFF,
            "vB": (val >> 8) & 0xFF,
            "vC": (val >> 16) & 0xFF,
            "State": (val >> 24) & 0x07,
            "RuinScore": (val >> 27) & 0x07
        }

    def load_environment_entropy(self, excel_path):
        """日本のExcel詳細設計書（方眼紙、結合セル、帳票形式）に特化した読み込み"""
        if not os.path.exists(excel_path):
            return False
            
        try:
            # 値のみを読み込む (数式ではなく評価結果を取得)
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            
            meta_data = f"【ブック構成情報】\nこのExcelファイルには以下のシートが含まれています: {', '.join(wb.sheetnames)}"
            self._add_sector("workbook_meta", meta_data)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 【日本のExcel対策1】結合セル（マージセル）の値を全対象セルに展開し、文脈の欠落を防ぐ
                merged_values = {}
                for merged_range in ws.merged_cells.ranges:
                    min_col, min_row, max_col, max_row = merged_range.bounds
                    top_left_val = ws.cell(row=min_row, column=min_col).value
                    if top_left_val is not None:
                        # セル内改行（Alt+Enter）は文脈が切れないようスペースに置換
                        val_str = str(top_left_val).replace('\n', ' ').replace('\r', '').strip()
                        if val_str:
                            for r in range(min_row, max_row + 1):
                                for c in range(min_col, max_col + 1):
                                    merged_values[(r, c)] = val_str
                
                chunk_lines = []
                chunk_length = 0
                chunk_idx = 0
                OVERLAP_SIZE = 2 
                
                # 【学習力向上】2Dプロジェクション用キャッシュ (上方ヘッダーの自動推論)
                col_headers = {}
                
                # 行ごとに走査
                for row_idx, row in enumerate(ws.iter_rows(), start=1):
                    row_data = []
                    last_val = None
                    
                    for col_idx, cell in enumerate(row, start=1):
                        # 結合セルの値があれば優先、なければ通常のセル値
                        if (row_idx, col_idx) in merged_values:
                            val_str = merged_values[(row_idx, col_idx)]
                        else:
                            val = cell.value
                            if val is not None:
                                val_str = str(val).replace('\n', ' ').replace('\r', '').strip()
                            else:
                                val_str = ""
                                
                        if val_str:
                            # 【日本のExcel対策2】方眼紙Excelによる同一テキストの連続出現（ノイズ）を排除
                            if val_str != last_val:
                                # 【日本のExcel対策4】2Dプロジェクション (上方の初出テキストをヘッダーとして結合)
                                if col_idx not in col_headers:
                                    col_headers[col_idx] = val_str
                                    row_data.append(val_str)
                                else:
                                    header = col_headers[col_idx]
                                    # ヘッダーと値が異なり、かつヘッダーが短め（20文字以内）ならキー・バリューとして採用
                                    if header != val_str and len(header) <= 20:
                                        row_data.append(f"[{header}] {val_str}")
                                    else:
                                        row_data.append(val_str)
                                last_val = val_str
                    
                    if not row_data:
                        continue
                        
                    # 【日本のExcel対策3】項目と値が「左から右へ」並ぶ関係性を '|' で繋ぐ
                    line_text = f"- 行{row_idx} => " + " | ".join(row_data)
                    
                    if chunk_length + len(line_text) > 800 and len(chunk_lines) > 0:
                        chunk_text = "\n".join(chunk_lines)
                        self._add_sector(f"{sheet_name}_part{chunk_idx}", f"【シート: {sheet_name}】\n{chunk_text}")
                        
                        overlap_lines = chunk_lines[-OVERLAP_SIZE:] if len(chunk_lines) > OVERLAP_SIZE else []
                        chunk_lines = overlap_lines + [line_text]
                        chunk_length = sum(len(l) for l in overlap_lines) + len(line_text)
                        chunk_idx += 1
                    else:
                        chunk_lines.append(line_text)
                        chunk_length += len(line_text)
                        
                if len(chunk_lines) > 0:
                    chunk_text = "\n".join(chunk_lines)
                    self._add_sector(f"{sheet_name}_part{chunk_idx}", f"【シート: {sheet_name}】\n{chunk_text}")
                    
            wb.close()
            return True
        except Exception as e:
            print(f"[内部エラー] エントロピー収穫中に異常発生: {e}")
            return False

    def _add_sector(self, sector_id, text_data):
        if self.active_sector_count >= self.max_sectors:
            return
            
        idx = self.active_sector_count
        self.ucdf_sectors.append({"id": sector_id, "data": text_data})
        
        bigrams = [text_data[i:i+2] for i in range(len(text_data) - 1)] if len(text_data) > 1 else [text_data]
        unique_bigrams = list(set(bigrams))
        
        if bigrams:
            hashes = [self._hash_bigram(bg) for bg in bigrams]
            hash_tensor = torch.tensor(hashes, dtype=torch.long, device=self.device)
            counts = torch.bincount(hash_tensor)
            non_zero_hashes = torch.nonzero(counts).squeeze(-1)
            self.feature_matrix[idx, non_zero_hashes] += counts[non_zero_hashes].to(torch.int16)
            
            unique_hashes = [self._hash_bigram(bg) for bg in unique_bigrams]
            unique_hash_tensor = torch.tensor(unique_hashes, dtype=torch.long, device=self.device)
            self.document_frequency[unique_hash_tensor] += 1
            
        self.pack_state(idx, va=1, vb=0, vc=0, state=1, ruin=0)
        
        # 抽出完了時に特徴量の総和（セクター長）を一度だけ計算してキャッシュしておく
        self.sector_lengths[idx] = torch.sum(self.feature_matrix[idx].to(torch.int32)) + 1
        
        self.active_sector_count += 1

    def calculate_cone_of_influence(self, user_query):
        if self.active_sector_count == 0:
            return "該当する参考資料はありません。"

        bigrams = [user_query[i:i+2] for i in range(len(user_query) - 1)] if len(user_query) > 1 else [user_query]
        unique_bigrams = list(set(bigrams))
        
        if not unique_bigrams:
            return "該当する参考資料はありません。"
            
        hashes = [self._hash_bigram(bg) for bg in unique_bigrams]
        hash_tensor = torch.tensor(hashes, dtype=torch.long, device=self.device)
        
        df_tensor = self.document_frequency[hash_tensor]
        idf_weights = (self.active_sector_count // (df_tensor + 1)) + 1
        
        relevant_features = self.feature_matrix[:self.active_sector_count, hash_tensor].to(torch.int32)
        base_scores = torch.mv(relevant_features, idf_weights.to(torch.int32))
        
        lengths = self.sector_lengths[:self.active_sector_count]
        base_scores_normalized = (base_scores * 100) // lengths
        
        states = self.state_buffer[:self.active_sector_count]
        va_values = states & 0xFF 
        
        final_scores = base_scores_normalized + va_values
        
        top_k = min(self.sector_capacity, self.active_sector_count)
        best_scores, best_indices = torch.topk(final_scores, top_k)
        
        # 【対策】動的カットオフの閾値を引き上げ（20% -> 40%）、ノイズ資料を厳格に排除
        if top_k > 0:
            max_score = best_scores[0].item()
            threshold = max_score * 0.4 
        else:
            threshold = 0

        cone_data = []
        for i in range(top_k):
            idx = best_indices[i].item()
            score = best_scores[i].item()
            if score > threshold and score > 0: 
                cone_data.append(self.ucdf_sectors[idx]["data"])
                
                s = self.unpack_state(idx)
                # アクセスされたセクターは重要度(vA)とアクセス数(vB)を回復・プロモート
                new_va = min(s["vA"] + 2, 255)
                new_vb = min(s["vB"] + 1, 255)
                self.pack_state(idx, va=new_va, vb=new_vb, vc=0, state=s["State"], ruin=s["RuinScore"])
                
        # 【対策】関連データがない場合は無理にフォールバックさせず「該当なし」を返す
        if not cone_data:
            return "該当する参考資料はありません。"
            
        return "\n".join(cone_data)

    def apply_vif_entropy_decay(self):
        """【アルゴリズム深化】VIF(動態遷移モデル)における忘却・減衰のテンソル一括演算
           ハードウェアの限界駆動により、ループを一切使わずに全セクターの状態を一瞬で更新する
        """
        if self.active_sector_count == 0:
            return
            
        states = self.state_buffer[:self.active_sector_count]
        
        va = states & 0xFF
        vb = (states >> 8) & 0xFF
        vc = (states >> 16) & 0xFF
        st = (states >> 24) & 0x07
        ruin = (states >> 27) & 0x07
        
        # 毎ターン、減衰値(vC)を増加
        vc = torch.clamp(vc + 1, max=255)
        
        # vCが閾値(3)を超えたら重要度(vA)を1減らし、vCをリセットする完全決定論的ビット演算
        decay_mask = (vc > 3).to(torch.int32)
        va = torch.clamp(va - decay_mask, min=1)
        vc = torch.where(decay_mask > 0, torch.zeros_like(vc), vc)
        
        # アクセスがなく重要度が最低のセクターはRuin(崩壊)スコアを蓄積
        ruin_mask = ((va == 1) & (vb == 0)).to(torch.int32)
        ruin = torch.clamp(ruin + ruin_mask, max=7)
        
        # 再パッキングしてバッファを一括上書き (CPU/GPUレイヤーでO(1)に近い速度)
        packed = (va | (vb << 8) | (vc << 16) | (st << 24) | (ruin << 27))
        self.state_buffer[:self.active_sector_count] = packed

    def enforce_phase_transition(self, user_text, ai_text):
        self.conversation_turns.append({"user": user_text, "ai": ai_text})
        if len(self.conversation_turns) > 3:
            self.conversation_turns.pop(0)
            self.probability_cloud_summary = "（注意：会話が長くなったため、一部の古い文脈は圧縮されました）"
            
        # ターン終了時にVIF減衰を一斉適用
        self.apply_vif_entropy_decay()


# =====================================================================
# UIレイヤー
# =====================================================================

def print_system_message(text):
    print(f"\n[システム] {text}")

def find_excel_file(filename):
    """3段階の多重検索ロジックで確実にファイルを探す"""
    path1 = os.path.abspath(filename)
    path2 = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    path3 = os.path.join(os.getcwd(), filename)
    
    for p in [path1, path2, path3]:
        if os.path.exists(p):
            return p
    return None

def main():
    os.system('cls' if os.name == 'nt' else 'clear')
    
    print("========================================")
    print("   社内資料 学習・作成支援アシスタント")
    print("========================================")
    
    engine = PyUCDF_ContextEngine(sector_capacity=5)
    excel_name = "design_docs.xlsx"
    
    print_system_message(f"社内資料（{excel_name}）を探しています...")
    
    excel_path = find_excel_file(excel_name)
    
    if excel_path:
        print_system_message(f"発見: {excel_path}")
        if engine.load_environment_entropy(excel_path):
            print_system_message("資料のベクトル化（2Dプロジェクション学習）、およびビットパッキングが完了しました。")
        else:
            print_system_message("ファイルの読み込み中にエラーが発生しました。")
    else:
        print_system_message(f"エラー: '{excel_name}' が見つかりませんでした。")
        engine._add_sector("dummy", "学習データなしで起動します。語尾は『〜とする』で統一します。")

    print_system_message("AIの知能をロードしています...")
    
    try:
        model_id = "Qwen/Qwen2.5-1.5B-Instruct"
        tokenizer = AutoTokenizer.from_pretrained(model_id)
        
        device = "cuda" if torch.cuda.is_available() else "cpu"
        
        if device == "cuda" and torch.cuda.is_bf16_supported():
            dtype = torch.bfloat16
        else:
            dtype = torch.float16 if device == "cuda" else torch.float32
        
        model = AutoModelForCausalLM.from_pretrained(
            model_id, 
            torch_dtype=dtype,
            low_cpu_mem_usage=True,
            attn_implementation="sdpa"
        ).to(device)
        
        print_system_message(f"準備完了 (使用デバイス: {device}, 推論精度: {dtype})")
        
    except Exception as e:
        print_system_message(f"AI読み込み失敗: {e}")
        return

    while True:
        try:
            user_input = input("\nあなた: ")
            if user_input.lower() in ['exit', 'quit', '終了']:
                break
            if not user_input.strip():
                continue

            print_system_message("参照データを抽出中...\n")
            
            reference_data = engine.calculate_cone_of_influence(user_input)
            
            # 【対策】システムプロンプトに禁止事項を強く明記し、一般知識の補完を防ぐ
            system_prompt = (
                "あなたは社内資料に基づいて的確かつ論理的に回答する優秀なAIアシスタントです。\n"
                "提供された【参考資料】の内容のみを事実とし、推測で捏造してはいけません。\n"
                "【重要】ユーザーの質問に対して、参考資料に関係のない情報が含まれている場合は、その部分には触れず、資料にある内容だけで回答してください。\n"
                "資料に全く記載がない場合は、あなたの持っている一般的な知識で答えず、必ず「資料に記載がありません」と明言してください。\n\n"
                "回答は必ず以下のフォーマットで出力してください。\n"
                "【結論】（質問に対する簡潔な直接の答え）\n"
                "【詳細】（参考資料から読み取れる具体的なデータや背景）"
            )
            if engine.probability_cloud_summary:
                system_prompt += f"\n{engine.probability_cloud_summary}"

            messages = [{"role": "system", "content": system_prompt}]
            for turn in engine.conversation_turns:
                messages.append({"role": "user", "content": turn["user"]})
                messages.append({"role": "assistant", "content": turn["ai"]})
            messages.append({"role": "user", "content": f"【参考資料】\n{reference_data}\n\n【質問】\n{user_input}"})
            
            prompt = tokenizer.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
            inputs = tokenizer(prompt, return_tensors="pt", add_special_tokens=False).to(device)
            streamer = TextStreamer(tokenizer, skip_prompt=True, skip_special_tokens=True)
            
            print("="*50 + "\n【作成支援AI】")
            with torch.inference_mode():
                outputs = model.generate(
                    **inputs,
                    max_new_tokens=2048, 
                    do_sample=False, 
                    repetition_penalty=1.1,
                    pad_token_id=tokenizer.eos_token_id,
                    streamer=streamer,
                    use_cache=True
                )
            print("\n" + "="*50)
            
            answer_tokens = outputs[0][inputs['input_ids'].shape[1]:]
            answer = tokenizer.decode(answer_tokens, skip_special_tokens=True).strip()
            engine.enforce_phase_transition(user_text=user_input, ai_text=answer)

        except KeyboardInterrupt:
            break

if __name__ == "__main__":
    main()